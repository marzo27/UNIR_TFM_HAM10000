"""
╔══════════════════════════════════════════════════════════════════════╗
║   SVM CLASSIFIER — HAM10000 Skin Lesion Dataset                      ║
║   Versión: 2.0  |  Línea base para comparación de modelos            ║
║   Genera: metricas.xlsx  con 7 hojas + 8 gráficos + JSON             ║
╚══════════════════════════════════════════════════════════════════════╝

Métricas capturadas (todas guardadas en metricas.xlsx):
  - Accuracy, F1 (macro/weighted/micro), Precision, Recall por clase
  - AUC-ROC (por clase + macro/weighted, one-vs-rest)
  - Average Precision / Curvas Precision-Recall por clase
  - Cohen's Kappa, Matthews Correlation Coefficient (MCC)
  - Validación cruzada 5-fold (accuracy + F1 macro)
  - Tiempos: carga, entrenamiento, inferencia total e inferencia/imagen
  - Análisis PCA: componentes, varianza explicada
  - Distribución de clases: original, train, test, post-SMOTE
  - Matriz de confusión (absoluta + normalizada)
  - Configuración completa del modelo

Hojas Excel generadas:
  1. RESUMEN              — métricas globales + evaluación semáforo
  2. MÉTRICAS_POR_CLASE   — detalle por las 7 clases HAM10000
  3. MATRIZ_CONFUSIÓN     — absoluta + normalizada + heatmap
  4. CURVAS_ROC_PR        — AUC por clase + imágenes de curvas
  5. VALIDACIÓN_CRUZADA   — scores fold a fold + estadísticos
  6. CONFIGURACIÓN        — todos los hiperparámetros y parámetros
  7. COMPARACIÓN_MODELOS  — tabla lista para añadir futuros modelos
"""

# ─────────────────────────────────────────────────────────────────────
# IMPORTACIONES
# ─────────────────────────────────────────────────────────────────────
import os
import sys
import time
import json
import warnings
import datetime
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')  # no requiere display
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import seaborn as sns
from PIL import Image
from tqdm import tqdm
from pathlib import Path

# Scikit-learn
from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
from sklearn.decomposition import PCA
from sklearn.svm import SVC
from sklearn.metrics import (
    classification_report, confusion_matrix,
    roc_auc_score, roc_curve, auc,
    precision_recall_curve, average_precision_score,
    cohen_kappa_score, matthews_corrcoef,
    accuracy_score, f1_score, precision_score, recall_score,
    balanced_accuracy_score
)
from sklearn.preprocessing import StandardScaler, label_binarize
from imblearn.over_sampling import SMOTE
import joblib

# Excel
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────────────────────────────
# ██  CONFIGURACIÓN — AJUSTA AQUÍ SI ES NECESARIO
# ─────────────────────────────────────────────────────────────────────
BASE_PATH   = Path(r'D:\TFM\HAM10000')   # raíz del dataset
OUTPUT_DIR  = Path(r'D:\TFM\Resultados')   # dónde se guardan los outputs
EXCEL_PATH  = OUTPUT_DIR / 'metricas.xlsx'
JSON_PATH   = OUTPUT_DIR / 'metricas_baseline.json'
MODEL_PATH  = OUTPUT_DIR / 'svm_model.joblib'
PLOT_DIR    = OUTPUT_DIR / 'plots_svm'

IMG_SIZE       = (64, 64)    # resolución de entrada (ancho × alto)
TEST_SIZE      = 0.20        # 20% para test
RANDOM_STATE   = 42
N_CV_FOLDS     = 5           # folds en validación cruzada
PCA_VARIANCE   = 0.95        # varianza a retener con PCA
MODEL_NAME     = 'SVM_RBF_PCA95_SMOTE'

# Umbrales para semáforo verde/amarillo/rojo en Excel: (good, mid)
# Referencia: ISIC 2018 Challenge Task 3 benchmarks + literatura dermatología IA
# Ajusta si tienes un requisito clínico o paper de referencia distinto.
THRESHOLDS = {
    "accuracy"           : (0.80, 0.65),
    "balanced_accuracy"  : (0.75, 0.60),
    "f1_macro"           : (0.70, 0.55),
    "f1_weighted"        : (0.80, 0.65),
    "f1_micro"           : (0.80, 0.65),
    "precision_macro"    : (0.70, 0.55),
    "precision_weighted" : (0.80, 0.65),
    "recall_macro"       : (0.70, 0.55),
    "recall_weighted"    : (0.80, 0.65),
    "auc_macro"          : (0.85, 0.75),
    "auc_weighted"       : (0.85, 0.75),
    "kappa"              : (0.60, 0.40),
    "mcc"                : (0.50, 0.30),
    "cv_acc"             : (0.75, 0.60),
    "cv_f1"              : (0.65, 0.50),
}

# Nombres completos de las 7 clases HAM10000
CLASS_NAMES = {
    'akiec': 'Actinic keratoses / Intraepithelial carcinoma',
    'bcc'  : 'Basal cell carcinoma',
    'bkl'  : 'Benign keratosis-like lesions',
    'df'   : 'Dermatofibroma',
    'mel'  : 'Melanoma',
    'nv'   : 'Melanocytic nevi',
    'vasc' : 'Vascular lesions',
}

TIMESTAMP = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
RUN_ID    = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')

# ─────────────────────────────────────────────────────────────────────
# PALETA Y ESTILOS PARA GRÁFICOS
# ─────────────────────────────────────────────────────────────────────
sns.set_style('whitegrid')
PALETTE    = sns.color_palette('husl', 7)
CLASSES_7  = list(CLASS_NAMES.keys())  # orden canónico; se ajusta al dataset real

# ─────────────────────────────────────────────────────────────────────
# HELPERS DE ESTILO PARA EXCEL
# ─────────────────────────────────────────────────────────────────────
FILL_DARK_BLUE   = PatternFill("solid", fgColor="1F497D")
FILL_MID_BLUE    = PatternFill("solid", fgColor="2E74B5")
FILL_LIGHT_BLUE  = PatternFill("solid", fgColor="BDD7EE")
FILL_ALT         = PatternFill("solid", fgColor="EBF3FB")
FILL_GREEN       = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW      = PatternFill("solid", fgColor="FFEB9C")
FILL_RED         = PatternFill("solid", fgColor="FFC7CE")
FILL_HEADER_COL  = PatternFill("solid", fgColor="4472C4")
FILL_WHITE       = PatternFill("solid", fgColor="FFFFFF")
FILL_GRAY        = PatternFill("solid", fgColor="F2F2F2")

FONT_WHITE_BOLD  = Font(bold=True, color="FFFFFF",  name="Calibri", size=11)
FONT_DARK_BOLD   = Font(bold=True, color="1F497D",  name="Calibri", size=11)
FONT_NORMAL      = Font(name="Calibri", size=10)
FONT_SMALL       = Font(name="Calibri", size=9)
FONT_TITLE       = Font(bold=True, color="FFFFFF",  name="Calibri", size=14)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

_side = Side(style='thin', color='BFBFBF')
BORDER_THIN = Border(left=_side, right=_side, top=_side, bottom=_side)

def _cell(ws, row, col, value=None, fill=None, font=None, align=None, border=None, fmt=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if fill   is not None: c.fill      = fill
    if font   is not None: c.font      = font
    if align  is not None: c.alignment = align
    if border is not None: c.border    = border
    if fmt    is not None: c.number_format = fmt
    return c

def title_row(ws, row, text, col_start, col_end, fill=FILL_DARK_BLUE, height=28):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    _cell(ws, row, col_start, text, fill=fill, font=FONT_TITLE, align=ALIGN_CENTER)
    ws.row_dimensions[row].height = height

def subheader_row(ws, row, text, col_start, col_end, fill=FILL_MID_BLUE):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    _cell(ws, row, col_start, text, fill=fill,
          font=Font(bold=True, color="FFFFFF", name="Calibri", size=11),
          align=ALIGN_CENTER)
    ws.row_dimensions[row].height = 22

def header_cells(ws, row, headers, start_col=1, fill=FILL_HEADER_COL):
    for i, h in enumerate(headers):
        _cell(ws, row, start_col+i, h,
              fill=fill, font=FONT_WHITE_BOLD,
              align=ALIGN_CENTER, border=BORDER_THIN)

def data_row(ws, row, values, start_col=1, alt=False):
    fill = FILL_ALT if alt else FILL_WHITE
    for i, v in enumerate(values):
        _cell(ws, row, start_col+i, v,
              fill=fill, font=FONT_NORMAL,
              align=ALIGN_CENTER, border=BORDER_THIN)

def set_col_widths(ws, widths):
    """widths: dict {col_letter_or_int: width}"""
    for col, w in widths.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w

def embed_image(ws, path, anchor, w=None, h=None):
    img = XLImage(str(path))
    if w: img.width  = w
    if h: img.height = h
    ws.add_image(img, anchor)

def metric_color(val, good=0.80, mid=0.65):
    if val >= good: return FILL_GREEN,  Font(bold=True, color="375623", name="Calibri", size=10)
    if val >= mid:  return FILL_YELLOW, Font(bold=True, color="9C5700", name="Calibri", size=10)
    return FILL_RED,    Font(bold=True, color="9C0006", name="Calibri", size=10)

def save_fig(fig, name):
    PLOT_DIR.mkdir(parents=True, exist_ok=True)
    path = PLOT_DIR / f"{name}.png"
    fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    return path


# ═════════════════════════════════════════════════════════════════════
# PASO 1 — DETECCIÓN AUTOMÁTICA DE ESTRUCTURA DEL DATASET
# ═════════════════════════════════════════════════════════════════════
print("=" * 65)
print("  SVM CLASSIFIER — HAM10000  |  Línea base de métricas")
print(f"  Ejecución: {TIMESTAMP}")
print("=" * 65)
print("\n[1/9] Detectando estructura del dataset...")

def detectar_dataset(base_path: Path):
    """
    Busca recursivamente el CSV de metadata y las carpetas de imágenes.
    Compatible con la estructura original de Kaggle (una o dos partes).
    """
    base_path = Path(base_path)
    if not base_path.exists():
        raise FileNotFoundError(f"La ruta base no existe: {base_path}")

    # --- CSV de metadata ---
    csv_found = (
        list(base_path.rglob('HAM10000_metadata.csv')) +
        list(base_path.rglob('*metadata*.csv'))
    )
    if not csv_found:
        raise FileNotFoundError(
            f"No se encontró 'HAM10000_metadata.csv' dentro de {base_path}\n"
            "Asegúrate de que el CSV esté en D:\\TFM\\HAM10000 o en una subcarpeta."
        )
    metadata_csv = csv_found[0]
    print(f"  ✓ Metadata : {metadata_csv}")

    # --- Carpetas de imágenes (deben contener .jpg) ---
    img_dirs = []
    for d in sorted(base_path.rglob('*')):
        if d.is_dir() and len(list(d.glob('*.jpg'))) > 50:
            img_dirs.append(d)
    if not img_dirs:
        raise FileNotFoundError(
            f"No se encontraron carpetas con imágenes .jpg dentro de {base_path}\n"
            "Estructura esperada:\n"
            "  D:\\TFM\\HAM10000\\HAM10000_images\\*.jpg\n"
            "  o bien dos partes:\n"
            "  D:\\TFM\\HAM10000\\HAM10000_images_part_1\\*.jpg\n"
            "  D:\\TFM\\HAM10000\\HAM10000_images_part_2\\*.jpg"
        )
    total_imgs = sum(len(list(d.glob('*.jpg'))) for d in img_dirs)
    print(f"  ✓ Carpetas : {[str(d.name) for d in img_dirs]}")
    print(f"  ✓ Imágenes : {total_imgs} archivos .jpg")
    return metadata_csv, img_dirs

metadata_csv, img_dirs = detectar_dataset(BASE_PATH)
df = pd.read_csv(metadata_csv)
print(f"  ✓ Metadata : {len(df)} filas  |  columnas: {list(df.columns)}")


# ═════════════════════════════════════════════════════════════════════
# PASO 2 — CARGA Y APLANADO DE IMÁGENES
# ═════════════════════════════════════════════════════════════════════
print(f"\n[2/9] Cargando imágenes a {IMG_SIZE[0]}×{IMG_SIZE[1]} px (RGB)...")

def cargar_imagenes(df, img_dirs, size):
    # Índice rápido: stem → Path
    lookup = {}
    for d in img_dirs:
        for f in d.glob('*.jpg'):
            lookup[f.stem] = f

    missing = 0
    X = []
    for img_id in tqdm(df['image_id'], desc="  Cargando", ncols=70):
        if img_id in lookup:
            try:
                arr = np.array(
                    Image.open(lookup[img_id]).convert('RGB').resize(size)
                ).flatten()
                X.append(arr)
            except Exception:
                X.append(np.zeros(size[0] * size[1] * 3, dtype=np.uint8))
                missing += 1
        else:
            X.append(np.zeros(size[0] * size[1] * 3, dtype=np.uint8))
            missing += 1

    if missing:
        print(f"  ⚠  {missing} imágenes no encontradas → reemplazadas con ceros")
    return np.array(X, dtype=np.float32)

t_load_start = time.time()
X = cargar_imagenes(df, img_dirs, IMG_SIZE)
t_load = time.time() - t_load_start
y = df['dx'].values

classes = sorted(np.unique(y))   # orden alfabético real del dataset
n_classes = len(classes)
color_map = {cls: PALETTE[i % len(PALETTE)] for i, cls in enumerate(classes)}

print(f"  ✓ X shape  : {X.shape}  ({X.nbytes / 1e6:.1f} MB)")
print(f"  ✓ Clases   : {classes}")
print(f"  ✓ Dist.    : {dict(zip(*np.unique(y, return_counts=True)))}")
print(f"  ✓ Tiempo   : {t_load:.1f} s")


# ═════════════════════════════════════════════════════════════════════
# PASO 3 — DIVISIÓN TRAIN / TEST
# ═════════════════════════════════════════════════════════════════════
print(f"\n[3/9] Dividiendo datos (train {int((1-TEST_SIZE)*100)}% / test {int(TEST_SIZE*100)}%)...")

X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=TEST_SIZE, random_state=RANDOM_STATE, stratify=y
)
dist_train_raw = dict(zip(*np.unique(y_train, return_counts=True)))
dist_test_raw  = dict(zip(*np.unique(y_test,  return_counts=True)))
print(f"  ✓ Train : {len(y_train)} muestras")
print(f"  ✓ Test  : {len(y_test)}  muestras")


# ═════════════════════════════════════════════════════════════════════
# PASO 4 — ESTANDARIZACIÓN + PCA
# ═════════════════════════════════════════════════════════════════════
print(f"\n[4/9] Escalando (StandardScaler) + PCA ({PCA_VARIANCE*100:.0f}% varianza)...")

scaler = StandardScaler()
X_train_sc = scaler.fit_transform(X_train)
X_test_sc  = scaler.transform(X_test)

pca = PCA(n_components=PCA_VARIANCE, random_state=RANDOM_STATE)
X_train_pca = pca.fit_transform(X_train_sc)
X_test_pca  = pca.transform(X_test_sc)

n_components     = pca.n_components_
var_ratio        = pca.explained_variance_ratio_
var_cumsum       = np.cumsum(var_ratio)
var_total        = float(var_cumsum[-1])
reduction_pct    = (1 - n_components / X_train.shape[1]) * 100

print(f"  ✓ Features orig  : {X_train.shape[1]}")
print(f"  ✓ Componentes PCA: {n_components}  (reducción {reduction_pct:.1f}%)")
print(f"  ✓ Varianza total : {var_total*100:.2f}%")


# ═════════════════════════════════════════════════════════════════════
# PASO 5 — SMOTE
# ═════════════════════════════════════════════════════════════════════
print("\n[5/9] Aplicando SMOTE (solo en train)...")

smote = SMOTE(random_state=RANDOM_STATE)
X_train_res, y_train_res = smote.fit_resample(X_train_pca, y_train)
dist_smote = dict(zip(*np.unique(y_train_res, return_counts=True)))

print(f"  ✓ Muestras post-SMOTE: {len(y_train_res)}")
print(f"  ✓ Distribución      : {dist_smote}")


# ═════════════════════════════════════════════════════════════════════
# PASO 6 — ENTRENAMIENTO SVM
# ═════════════════════════════════════════════════════════════════════
print("\n[6/9] Entrenando SVM con kernel RBF (puede tardar varios minutos)...")

t_train_start = time.time()
svm = SVC(
    kernel='rbf',
    C=1.0,
    gamma='scale',
    class_weight='balanced',
    probability=True,          # necesario para AUC / Platt scaling
    random_state=RANDOM_STATE,
    decision_function_shape='ovr',
)
svm.fit(X_train_res, y_train_res)
t_train = time.time() - t_train_start

print(f"  ✓ Entrenamiento completado en {t_train:.1f} s")
print(f"  ✓ Vectores de soporte: {svm.n_support_.sum()} total  {dict(zip(svm.classes_, svm.n_support_))}")

# Guardar modelo completo
joblib.dump({'model': svm, 'scaler': scaler, 'pca': pca,
             'classes': classes, 'img_size': IMG_SIZE}, MODEL_PATH)
print(f"  ✓ Modelo guardado en: {MODEL_PATH}")


# ═════════════════════════════════════════════════════════════════════
# PASO 7 — EVALUACIÓN COMPLETA
# ═════════════════════════════════════════════════════════════════════
print("\n[7/9] Evaluando modelo en test set...")

t_infer_start = time.time()
y_pred  = svm.predict(X_test_pca)
y_proba = svm.predict_proba(X_test_pca)   # shape (n_test, n_classes)
t_infer = time.time() - t_infer_start
t_per_img_ms = t_infer / len(y_test) * 1000

# --- Métricas globales ---
acc          = accuracy_score(y_test, y_pred)
bal_acc      = balanced_accuracy_score(y_test, y_pred)
f1_macro     = f1_score(y_test, y_pred, average='macro',    zero_division=0)
f1_weighted  = f1_score(y_test, y_pred, average='weighted', zero_division=0)
f1_micro     = f1_score(y_test, y_pred, average='micro',    zero_division=0)
prec_macro   = precision_score(y_test, y_pred, average='macro',    zero_division=0)
prec_w       = precision_score(y_test, y_pred, average='weighted', zero_division=0)
rec_macro    = recall_score(y_test, y_pred, average='macro',    zero_division=0)
rec_w        = recall_score(y_test, y_pred, average='weighted', zero_division=0)
kappa        = cohen_kappa_score(y_test, y_pred)
mcc          = matthews_corrcoef(y_test, y_pred)

# --- AUC (one-vs-rest) ---
y_bin = label_binarize(y_test, classes=classes)
auc_per_class = {}
for i, cls in enumerate(classes):
    try:
        auc_per_class[cls] = roc_auc_score(y_bin[:, i], y_proba[:, i])
    except Exception:
        auc_per_class[cls] = float('nan')
try:
    auc_macro    = roc_auc_score(y_bin, y_proba, average='macro',    multi_class='ovr')
    auc_weighted = roc_auc_score(y_bin, y_proba, average='weighted', multi_class='ovr')
except Exception:
    auc_macro = auc_weighted = float('nan')

# --- Average Precision por clase ---
ap_per_class = {}
for i, cls in enumerate(classes):
    try:
        ap_per_class[cls] = average_precision_score(y_bin[:, i], y_proba[:, i])
    except Exception:
        ap_per_class[cls] = float('nan')

# --- Reporte sklearn ---
report_dict = classification_report(
    y_test, y_pred, target_names=classes, output_dict=True, zero_division=0
)

# --- Matrices de confusión ---
cm      = confusion_matrix(y_test, y_pred, labels=classes)
cm_norm = cm.astype(float) / cm.sum(axis=1, keepdims=True)

print(f"  ✓ Accuracy         : {acc:.4f}")
print(f"  ✓ Balanced Accuracy: {bal_acc:.4f}")
print(f"  ✓ F1 Macro         : {f1_macro:.4f}")
print(f"  ✓ F1 Weighted      : {f1_weighted:.4f}")
print(f"  ✓ AUC Macro (OvR)  : {auc_macro:.4f}")
print(f"  ✓ Cohen's Kappa    : {kappa:.4f}")
print(f"  ✓ MCC              : {mcc:.4f}")


# ═════════════════════════════════════════════════════════════════════
# PASO 8 — VALIDACIÓN CRUZADA (sobre datos PCA, sin SMOTE)
# ═════════════════════════════════════════════════════════════════════
print(f"\n[8/9] Validación cruzada {N_CV_FOLDS}-fold (puede tardar)...")

skf = StratifiedKFold(n_splits=N_CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)
_svm_cv = SVC(kernel='rbf', class_weight='balanced', random_state=RANDOM_STATE)

cv_acc = cross_val_score(_svm_cv, X_train_pca, y_train,
                         cv=skf, scoring='accuracy', n_jobs=-1)
cv_f1  = cross_val_score(_svm_cv, X_train_pca, y_train,
                         cv=skf, scoring='f1_macro',  n_jobs=-1)

print(f"  ✓ CV Accuracy: {cv_acc.mean():.4f} ± {cv_acc.std():.4f}")
print(f"  ✓ CV F1 Macro: {cv_f1.mean():.4f}  ± {cv_f1.std():.4f}")


# ═════════════════════════════════════════════════════════════════════
# PASO 9 — GRÁFICOS
# ═════════════════════════════════════════════════════════════════════
print("\n[9/9] Generando gráficos y Excel...")
PLOT_DIR.mkdir(parents=True, exist_ok=True)


# ── Gráfico 1: Distribución de clases ──────────────────────────────
fig, axes = plt.subplots(1, 3, figsize=(18, 5))
fig.suptitle('Distribución de Clases — HAM10000', fontsize=14, fontweight='bold')
datasets = [
    (dist_train_raw, f'Train (n={len(y_train)})'),
    (dist_test_raw,  f'Test  (n={len(y_test)})'),
    (dist_smote,     f'Train + SMOTE (n={len(y_train_res)})'),
]
for ax, (dist, ttl) in zip(axes, datasets):
    cols = [color_map.get(c, 'steelblue') for c in classes]
    vals = [dist.get(c, 0) for c in classes]
    bars = ax.bar(classes, vals, color=cols, edgecolor='white', linewidth=0.8)
    ax.set_title(ttl, fontsize=11)
    ax.set_xlabel('Clase'); ax.set_ylabel('Muestras')
    ax.tick_params(axis='x', rotation=45)
    for b, v in zip(bars, vals):
        ax.text(b.get_x() + b.get_width()/2, b.get_height() + 5,
                str(v), ha='center', va='bottom', fontsize=8)
plt.tight_layout()
p_dist = save_fig(fig, '01_distribucion_clases')

# ── Gráfico 2: Análisis PCA ────────────────────────────────────────
fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
fig.suptitle('Análisis de Componentes Principales (PCA)', fontsize=14, fontweight='bold')

top_n = min(60, n_components)
ax1.bar(range(1, top_n+1), var_ratio[:top_n]*100, color='steelblue', alpha=0.7)
ax1.set_title(f'Varianza explicada por componente (top {top_n})')
ax1.set_xlabel('Componente'); ax1.set_ylabel('Varianza (%)')

ax2.plot(range(1, n_components+1), var_cumsum*100, 'b-', linewidth=2)
ax2.axhline(95, color='red', linestyle='--', label='95%')
ax2.fill_between(range(1, n_components+1), var_cumsum*100, alpha=0.15, color='steelblue')
ax2.set_title('Varianza acumulada')
ax2.set_xlabel('Núm. componentes'); ax2.set_ylabel('Varianza acumulada (%)')
ax2.legend(); ax2.grid(True, alpha=0.3)
plt.tight_layout()
p_pca = save_fig(fig, '02_analisis_pca')

# ── Gráfico 3: Matriz de confusión ────────────────────────────────
fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))
fig.suptitle('Matriz de Confusión — SVM RBF', fontsize=14, fontweight='bold')
sns.heatmap(cm,      annot=True, fmt='d',    cmap='Blues',
            xticklabels=classes, yticklabels=classes, ax=ax1,
            linewidths=0.5, linecolor='white')
ax1.set_title('Conteos absolutos'); ax1.set_ylabel('Real'); ax1.set_xlabel('Predicho')
sns.heatmap(cm_norm, annot=True, fmt='.2f',  cmap='Blues',
            xticklabels=classes, yticklabels=classes, ax=ax2,
            linewidths=0.5, linecolor='white')
ax2.set_title('Normalizada (recall por clase)'); ax2.set_ylabel('Real'); ax2.set_xlabel('Predicho')
plt.tight_layout()
p_cm = save_fig(fig, '03_matriz_confusion')

# ── Gráfico 4: Curvas ROC ─────────────────────────────────────────
ncols = 4
nrows = (n_classes + 1 + ncols - 1) // ncols  # +1 para el panel "todas"
fig, axes = plt.subplots(nrows, ncols, figsize=(ncols*5, nrows*4))
fig.suptitle('Curvas ROC por Clase (One-vs-Rest)', fontsize=14, fontweight='bold')
axes_flat = axes.flatten() if nrows > 1 else axes

for i, cls in enumerate(classes):
    fpr, tpr, _ = roc_curve(y_bin[:, i], y_proba[:, i])
    ax = axes_flat[i]
    ax.plot(fpr, tpr, color=PALETTE[i], lw=2, label=f'AUC = {auc_per_class[cls]:.3f}')
    ax.plot([0,1],[0,1],'k--', lw=0.8)
    ax.set_title(f'{cls}\n{CLASS_NAMES.get(cls,"")[:30]}', fontsize=9)
    ax.set_xlabel('FPR'); ax.set_ylabel('TPR')
    ax.legend(loc='lower right', fontsize=9)
    ax.grid(True, alpha=0.3)

# Panel "todas las clases"
all_ax = axes_flat[n_classes]
for i, cls in enumerate(classes):
    fpr, tpr, _ = roc_curve(y_bin[:, i], y_proba[:, i])
    all_ax.plot(fpr, tpr, color=PALETTE[i], lw=1.5,
                label=f'{cls} ({auc_per_class[cls]:.3f})')
all_ax.plot([0,1],[0,1],'k--', lw=0.8)
all_ax.set_title('Todas las clases\n(AUC macro={:.3f})'.format(auc_macro))
all_ax.legend(loc='lower right', fontsize=8)
all_ax.grid(True, alpha=0.3)

for j in range(n_classes+1, len(axes_flat)):
    axes_flat[j].set_visible(False)
plt.tight_layout()
p_roc = save_fig(fig, '04_curvas_roc')

# ── Gráfico 5: Curvas Precision-Recall ───────────────────────────
fig, axes = plt.subplots(nrows, ncols, figsize=(ncols*5, nrows*4))
fig.suptitle('Curvas Precision–Recall por Clase', fontsize=14, fontweight='bold')
axes_flat = axes.flatten() if nrows > 1 else axes

for i, cls in enumerate(classes):
    prec, rec, _ = precision_recall_curve(y_bin[:, i], y_proba[:, i])
    ap = ap_per_class[cls]
    ax = axes_flat[i]
    ax.plot(rec, prec, color=PALETTE[i], lw=2, label=f'AP = {ap:.3f}')
    baseline = y_bin[:, i].mean()
    ax.axhline(baseline, color='gray', linestyle='--', lw=0.8, label=f'Baseline={baseline:.2f}')
    ax.set_title(f'{cls}\n{CLASS_NAMES.get(cls,"")[:30]}', fontsize=9)
    ax.set_xlabel('Recall'); ax.set_ylabel('Precision')
    ax.legend(loc='upper right', fontsize=9); ax.grid(True, alpha=0.3)

all_ax = axes_flat[n_classes]
for i, cls in enumerate(classes):
    prec, rec, _ = precision_recall_curve(y_bin[:, i], y_proba[:, i])
    all_ax.plot(rec, prec, color=PALETTE[i], lw=1.5,
                label=f'{cls} ({ap_per_class[cls]:.3f})')
all_ax.set_title('Todas las clases'); all_ax.legend(loc='upper right', fontsize=8)
all_ax.grid(True, alpha=0.3)

for j in range(n_classes+1, len(axes_flat)):
    axes_flat[j].set_visible(False)
plt.tight_layout()
p_pr = save_fig(fig, '05_precision_recall')

# ── Gráfico 6: Métricas por clase (barras agrupadas) ──────────────
metrics_df = pd.DataFrame({
    'Precision': [report_dict[c]['precision'] for c in classes],
    'Recall'   : [report_dict[c]['recall']    for c in classes],
    'F1-Score' : [report_dict[c]['f1-score']  for c in classes],
    'AUC'      : [auc_per_class.get(c, 0)     for c in classes],
}, index=classes)

fig, axes = plt.subplots(2, 2, figsize=(16, 10))
fig.suptitle('Métricas por Clase — SVM RBF', fontsize=14, fontweight='bold')
axes_flat = axes.flatten()

for ax, metric in zip(axes_flat, ['Precision','Recall','F1-Score','AUC']):
    vals = metrics_df[metric]
    cols = [PALETTE[i] for i in range(len(classes))]
    bars = ax.bar(classes, vals, color=cols, edgecolor='white')
    ax.set_title(metric, fontsize=12, fontweight='bold')
    ax.set_ylim(0, 1.12)
    ax.set_ylabel(metric)
    ax.tick_params(axis='x', rotation=45)
    mu = vals.mean()
    ax.axhline(mu, color='red', linestyle='--', alpha=0.6,
               label=f'Media: {mu:.3f}')
    ax.legend(fontsize=9)
    for b, v in zip(bars, vals):
        ax.text(b.get_x()+b.get_width()/2, b.get_height()+0.01,
                f'{v:.3f}', ha='center', va='bottom', fontsize=8)
plt.tight_layout()
p_metric_class = save_fig(fig, '06_metricas_por_clase')

# ── Gráfico 7: Validación cruzada ─────────────────────────────────
fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
fig.suptitle(f'Validación Cruzada {N_CV_FOLDS}-Fold', fontsize=14, fontweight='bold')

for ax, scores, title, col in [
    (ax1, cv_acc, 'Accuracy',  'steelblue'),
    (ax2, cv_f1,  'F1 Macro',  'darkorange'),
]:
    folds = [f'F{i+1}' for i in range(N_CV_FOLDS)]
    bars = ax.bar(folds, scores, color=col, alpha=0.75, edgecolor='white')
    mu, sd = scores.mean(), scores.std()
    ax.axhline(mu, color='red', linestyle='--',
               label=f'μ={mu:.4f}  σ={sd:.4f}')
    ax.set_title(title); ax.set_ylabel(title)
    ax.set_ylim(max(0, mu-0.15), min(1, mu+0.15))
    ax.legend(fontsize=9)
    for b, v in zip(bars, scores):
        ax.text(b.get_x()+b.get_width()/2, b.get_height()+0.002,
                f'{v:.4f}', ha='center', va='bottom', fontsize=9)
plt.tight_layout()
p_cv = save_fig(fig, '07_validacion_cruzada')

# ── Gráfico 8: Radar / Spider chart de métricas globales ──────────
from matplotlib.patches import FancyArrowPatch

radar_metrics = {
    'Accuracy'   : acc,
    'Bal.Acc'    : bal_acc,
    'F1 Macro'   : f1_macro,
    'F1 Weighted': f1_weighted,
    'AUC Macro'  : auc_macro,
    'Kappa'      : kappa,
    'MCC'        : (mcc + 1) / 2,  # normalizar a [0,1]
    'CV Acc'     : cv_acc.mean(),
    'CV F1'      : cv_f1.mean(),
}
labels = list(radar_metrics.keys())
values = list(radar_metrics.values())
N = len(labels)
angles = np.linspace(0, 2*np.pi, N, endpoint=False).tolist()
angles += angles[:1]
vals_plot = values + values[:1]

fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))
fig.suptitle('Radar de Métricas Globales — SVM RBF\n(MCC normalizado a [0,1])',
             fontsize=13, fontweight='bold')
ax.plot(angles, vals_plot, 'b-o', linewidth=2, markersize=6)
ax.fill(angles, vals_plot, alpha=0.2, color='steelblue')
ax.set_thetagrids(np.degrees(angles[:-1]), labels, fontsize=10)
ax.set_ylim(0, 1)
for i, (a, v) in enumerate(zip(angles[:-1], values)):
    ax.text(a, v+0.05, f'{v:.3f}', ha='center', va='center', fontsize=9, color='navy')
ax.set_yticklabels(['0.2','0.4','0.6','0.8','1.0'], fontsize=8)
plt.tight_layout()
p_radar = save_fig(fig, '08_radar_metricas')

plot_paths = {
    'distribucion'   : p_dist,
    'pca'            : p_pca,
    'confusion'      : p_cm,
    'roc'            : p_roc,
    'pr'             : p_pr,
    'metricas_clase' : p_metric_class,
    'cv'             : p_cv,
    'radar'          : p_radar,
}
print(f"  ✓ {len(plot_paths)} gráficos guardados en: {PLOT_DIR}")


# ═════════════════════════════════════════════════════════════════════
# EXCEL — CONSTRUCCIÓN DEL WORKBOOK
# ═════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ──────────────────────────────────────────────────────────────────
# HOJA 1 : RESUMEN
# ──────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "RESUMEN"
ws1.sheet_view.showGridLines = False
set_col_widths(ws1, {1:32, 2:18, 3:20, 4:16})

title_row(ws1, 1, f"MÉTRICAS — {MODEL_NAME}  |  HAM10000", 1, 4)

# Bloque configuración
r = 3
subheader_row(ws1, r, "CONFIGURACIÓN DEL EXPERIMENTO", 1, 4); r += 1
header_cells(ws1, r, ["Parámetro", "Valor", "", ""], start_col=1)
ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4); r += 1

cfg_rows = [
    ("Fecha ejecución",          TIMESTAMP),
    ("Dataset",                  "HAM10000 Skin Lesion Dataset"),
    ("Total imágenes",           len(df)),
    ("Clases",                   ', '.join(classes)),
    ("Num. clases",              n_classes),
    ("Train / Test split",       f"{int((1-TEST_SIZE)*100)}% / {int(TEST_SIZE*100)}%"),
    ("Train muestras",           len(y_train)),
    ("Test muestras",            len(y_test)),
    ("Train post-SMOTE",         len(y_train_res)),
    ("Tamaño imagen",            f"{IMG_SIZE[0]}×{IMG_SIZE[1]} px RGB"),
    ("Features orig.",           int(X_train.shape[1])),
    ("Componentes PCA",          n_components),
    ("Varianza PCA capturada",   f"{var_total*100:.2f}%"),
    ("Tiempo carga imágenes",    f"{t_load:.1f} s"),
    ("Tiempo entrenamiento",     f"{t_train:.1f} s"),
    ("Tiempo inferencia total",  f"{t_infer:.3f} s"),
    ("Inferencia por imagen",    f"{t_per_img_ms:.3f} ms"),
]
for i, (k, v) in enumerate(cfg_rows):
    ws1.cell(row=r, column=1, value=k).font = Font(bold=True, name="Calibri", size=10)
    c2 = ws1.cell(row=r, column=2, value=str(v))
    c2.alignment = ALIGN_CENTER; c2.font = FONT_NORMAL
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    for col in range(1, 5):
        ws1.cell(row=r, column=col).border = BORDER_THIN
        if i % 2 == 0:
            ws1.cell(row=r, column=col).fill = FILL_ALT
    r += 1

# Bloque métricas globales
r += 1
subheader_row(ws1, r, "MÉTRICAS GLOBALES", 1, 4); r += 1
header_cells(ws1, r, ["Métrica", "Valor", "Umbral (bueno)", "Estado"])
r += 1

global_metrics = [
    ("Accuracy",                            acc,           *THRESHOLDS["accuracy"]),
    ("Balanced Accuracy",                   bal_acc,       *THRESHOLDS["balanced_accuracy"]),
    ("F1-Score Macro",                      f1_macro,      *THRESHOLDS["f1_macro"]),
    ("F1-Score Weighted",                   f1_weighted,   *THRESHOLDS["f1_weighted"]),
    ("F1-Score Micro",                      f1_micro,      *THRESHOLDS["f1_micro"]),
    ("Precision Macro",                     prec_macro,    *THRESHOLDS["precision_macro"]),
    ("Precision Weighted",                  prec_w,        *THRESHOLDS["precision_weighted"]),
    ("Recall Macro",                        rec_macro,     *THRESHOLDS["recall_macro"]),
    ("Recall Weighted",                     rec_w,         *THRESHOLDS["recall_weighted"]),
    ("AUC Macro (OvR)",                     auc_macro,     *THRESHOLDS["auc_macro"]),
    ("AUC Weighted (OvR)",                  auc_weighted,  *THRESHOLDS["auc_weighted"]),
    ("Cohen's Kappa",                       kappa,         *THRESHOLDS["kappa"]),
    ("Matthews CC (MCC)",                   mcc,           *THRESHOLDS["mcc"]),
    (f"CV Accuracy {N_CV_FOLDS}-fold (μ)",  cv_acc.mean(), *THRESHOLDS["cv_acc"]),
    (f"CV F1 Macro {N_CV_FOLDS}-fold (μ)",  cv_f1.mean(),  *THRESHOLDS["cv_f1"]),
    (f"CV Accuracy {N_CV_FOLDS}-fold (σ)",  cv_acc.std(),  None, None),
    (f"CV F1 Macro {N_CV_FOLDS}-fold (σ)",  cv_f1.std(),   None, None),
]
for i, (name, val, good, mid) in enumerate(global_metrics):
    fill_r = FILL_ALT if i % 2 == 0 else FILL_WHITE
    ws1.cell(row=r, column=1, value=name).font = Font(name="Calibri", size=10)
    ws1.cell(row=r, column=1).fill = fill_r
    ws1.cell(row=r, column=1).border = BORDER_THIN
    v_cell = ws1.cell(row=r, column=2, value=round(float(val), 4))
    v_cell.alignment = ALIGN_CENTER; v_cell.font = Font(bold=True, name="Calibri", size=10)
    v_cell.fill = fill_r; v_cell.border = BORDER_THIN
    if good is not None:
        ws1.cell(row=r, column=3, value=f">= {good}").alignment = ALIGN_CENTER
        ws1.cell(row=r, column=3).font = FONT_NORMAL
        ws1.cell(row=r, column=3).fill = fill_r
        ws1.cell(row=r, column=3).border = BORDER_THIN
        ef, ef_font = metric_color(val, good=good, mid=mid)
        label = "✓ Bueno" if val >= good else ("~ Aceptable" if val >= mid else "✗ Bajo")
        e_cell = ws1.cell(row=r, column=4, value=label)
        e_cell.fill = ef; e_cell.font = ef_font
        e_cell.alignment = ALIGN_CENTER; e_cell.border = BORDER_THIN
    else:
        for col in [3, 4]:
            ws1.cell(row=r, column=col, value="—").alignment = ALIGN_CENTER
            ws1.cell(row=r, column=col).fill = fill_r
            ws1.cell(row=r, column=col).border = BORDER_THIN
    r += 1

# Imágenes
r += 2
embed_image(ws1, p_radar,        f"A{r}", w=420, h=420); r += 27
embed_image(ws1, p_metric_class, f"A{r}", w=730, h=380); r += 26
embed_image(ws1, p_dist,         f"A{r}", w=730, h=220)


# ──────────────────────────────────────────────────────────────────
# HOJA 2 : MÉTRICAS POR CLASE
# ──────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("MÉTRICAS_POR_CLASE")
ws2.sheet_view.showGridLines = False
set_col_widths(ws2, {1:8, 2:38, 3:12, 4:12, 5:12, 6:10, 7:12, 8:12, 9:10})
title_row(ws2, 1, "MÉTRICAS DETALLADAS POR CLASE — HAM10000", 1, 9)

r = 3
header_cells(ws2, r, ["Clase","Nombre completo","Precision","Recall","F1-Score",
                       "Support","AUC (OvR)","Avg Prec.","% test"])
r += 1
total_test = len(y_test)
for i, cls in enumerate(classes):
    sup = int(report_dict[cls]['support'])
    row_vals = [
        cls,
        CLASS_NAMES.get(cls, cls),
        round(report_dict[cls]['precision'], 4),
        round(report_dict[cls]['recall'],    4),
        round(report_dict[cls]['f1-score'],  4),
        sup,
        round(auc_per_class.get(cls, 0), 4),
        round(ap_per_class.get(cls, 0),  4),
        round(sup / total_test * 100, 1),
    ]
    data_row(ws2, r, row_vals, alt=(i % 2 == 0))
    # Semáforo F1
    f1v = report_dict[cls]['f1-score']
    ef, ef_font = metric_color(f1v, good=0.70, mid=0.50)
    ws2.cell(row=r, column=5).fill = ef
    ws2.cell(row=r, column=5).font = Font(bold=True, name="Calibri", size=10,
                                          color=ef_font.color)
    r += 1

# Filas de promedios
for avg_key in ['macro avg', 'weighted avg']:
    ws2.cell(row=r, column=1, value=avg_key).font = Font(bold=True, name="Calibri", size=10)
    ws2.cell(row=r, column=2, value="—").alignment = ALIGN_CENTER
    for c, key in [(3,'precision'),(4,'recall'),(5,'f1-score')]:
        _cell(ws2, r, c, round(report_dict[avg_key][key], 4),
              fill=FILL_LIGHT_BLUE,
              font=Font(bold=True, name="Calibri", size=10),
              align=ALIGN_CENTER, border=BORDER_THIN)
    _cell(ws2, r, 6, int(report_dict[avg_key]['support']),
          fill=FILL_LIGHT_BLUE, font=Font(bold=True, name="Calibri"), align=ALIGN_CENTER, border=BORDER_THIN)
    if avg_key == 'macro avg':
        _cell(ws2, r, 7, round(auc_macro, 4),
              fill=FILL_LIGHT_BLUE, font=Font(bold=True, name="Calibri"), align=ALIGN_CENTER, border=BORDER_THIN)
    r += 1

r += 2
embed_image(ws2, p_roc, f"A{r}", w=760, h=400); r += 28
embed_image(ws2, p_pr,  f"A{r}", w=760, h=400)


# ──────────────────────────────────────────────────────────────────
# HOJA 3 : MATRIZ DE CONFUSIÓN
# ──────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("MATRIZ_CONFUSIÓN")
ws3.sheet_view.showGridLines = False
set_col_widths(ws3, {i+1: 12 for i in range(n_classes+2)})
title_row(ws3, 1, "MATRIZ DE CONFUSIÓN", 1, n_classes+2)

def write_cm_block(ws, start_row, cm_data, fmt, title_text, is_pct=False):
    ws.cell(row=start_row, column=1,
            value=title_text).font = Font(bold=True, color="1F497D", name="Calibri", size=11)
    r = start_row + 1
    # encabezado
    _cell(ws, r, 1, "Real \\ Pred",
          fill=FILL_DARK_BLUE, font=FONT_WHITE_BOLD, align=ALIGN_CENTER, border=BORDER_THIN)
    for c, cls in enumerate(classes, 2):
        _cell(ws, r, c, cls,
              fill=FILL_DARK_BLUE, font=FONT_WHITE_BOLD, align=ALIGN_CENTER, border=BORDER_THIN)
    r += 1
    max_val = cm_data.max()
    for i, row_cls in enumerate(classes):
        _cell(ws, r+i, 1, row_cls,
              fill=FILL_MID_BLUE, font=FONT_WHITE_BOLD, align=ALIGN_CENTER, border=BORDER_THIN)
        for j in range(n_classes):
            raw = cm_data[i, j]
            val = float(raw)
            cell = ws.cell(row=r+i, column=j+2)
            cell.value = round(val, 3) if is_pct else int(val)
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_THIN
            cell.number_format = fmt
            # Color gradient
            if i == j:
                intensity = int(55 + 180 * (val / max(max_val, 1e-9)))
                hex_g = f"{intensity:02X}"
                cell.fill = PatternFill("solid", fgColor=f"00{hex_g}00")
                cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            else:
                intensity = int(255 - 180 * (val / max(max_val, 1e-9)))
                hex_i = f"{intensity:02X}"
                cell.fill = PatternFill("solid", fgColor=f"FF{hex_i}{hex_i}")
                cell.font = FONT_NORMAL

r_block = 3
write_cm_block(ws3, r_block, cm,      "0",    "CONTEOS ABSOLUTOS")
r_block += n_classes + 4
write_cm_block(ws3, r_block, cm_norm, "0.0%", "NORMALIZADA — Recall por clase (diagonal = TPR)", is_pct=True)
r_block += n_classes + 4
embed_image(ws3, p_cm, f"A{r_block}", w=700, h=300)


# ──────────────────────────────────────────────────────────────────
# HOJA 4 : CURVAS ROC / PR
# ──────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("CURVAS_ROC_PR")
ws4.sheet_view.showGridLines = False
set_col_widths(ws4, {1:18, 2:20, 3:14, 4:14, 5:14})
title_row(ws4, 1, "CURVAS ROC y PRECISION–RECALL por Clase", 1, 5)

r = 3
subheader_row(ws4, r, "AUC-ROC y Average Precision", 1, 5); r += 1
header_cells(ws4, r, ["Clase","Nombre completo","AUC (OvR)","Avg Precision","Soporte"])
r += 1
for i, cls in enumerate(classes):
    auc_v = auc_per_class.get(cls, 0)
    ap_v  = ap_per_class.get(cls, 0)
    row_vals = [cls, CLASS_NAMES.get(cls, cls),
                round(auc_v, 4), round(ap_v, 4),
                int(report_dict[cls]['support'])]
    data_row(ws4, r, row_vals, alt=(i % 2 == 0))
    ef, ef_font = metric_color(auc_v, good=0.85, mid=0.70)
    ws4.cell(row=r, column=3).fill = ef
    ws4.cell(row=r, column=3).font = Font(bold=True, name="Calibri", size=10, color=ef_font.color)
    r += 1

_cell(ws4, r, 1, "MACRO", fill=FILL_LIGHT_BLUE,
      font=Font(bold=True, name="Calibri"), align=ALIGN_CENTER, border=BORDER_THIN)
_cell(ws4, r, 2, "Promedio no ponderado", fill=FILL_LIGHT_BLUE,
      font=Font(bold=True, name="Calibri"), align=ALIGN_CENTER, border=BORDER_THIN)
_cell(ws4, r, 3, round(auc_macro, 4), fill=FILL_LIGHT_BLUE,
      font=Font(bold=True, name="Calibri"), align=ALIGN_CENTER, border=BORDER_THIN)
r += 1
_cell(ws4, r, 1, "WEIGHTED", fill=FILL_LIGHT_BLUE,
      font=Font(bold=True, name="Calibri"), align=ALIGN_CENTER, border=BORDER_THIN)
_cell(ws4, r, 2, "Ponderado por soporte", fill=FILL_LIGHT_BLUE,
      font=Font(bold=True, name="Calibri"), align=ALIGN_CENTER, border=BORDER_THIN)
_cell(ws4, r, 3, round(auc_weighted, 4), fill=FILL_LIGHT_BLUE,
      font=Font(bold=True, name="Calibri"), align=ALIGN_CENTER, border=BORDER_THIN)
r += 3

embed_image(ws4, p_roc, f"A{r}", w=750, h=390); r += 28
embed_image(ws4, p_pr,  f"A{r}", w=750, h=390)


# ──────────────────────────────────────────────────────────────────
# HOJA 5 : VALIDACIÓN CRUZADA
# ──────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("VALIDACIÓN_CRUZADA")
ws5.sheet_view.showGridLines = False
set_col_widths(ws5, {1:14, 2:16, 3:16})
title_row(ws5, 1, f"VALIDACIÓN CRUZADA {N_CV_FOLDS}-FOLD (sobre datos PCA, sin SMOTE)", 1, 3)

r = 3
subheader_row(ws5, r, "Scores por fold", 1, 3); r += 1
header_cells(ws5, r, ["Fold", "Accuracy", "F1 Macro"]); r += 1
for i in range(N_CV_FOLDS):
    data_row(ws5, r, [f"Fold {i+1}", round(cv_acc[i], 4), round(cv_f1[i], 4)], alt=(i%2==0))
    r += 1

r += 1
subheader_row(ws5, r, "Estadísticos", 1, 3); r += 1
header_cells(ws5, r, ["Estadístico", "Accuracy", "F1 Macro"]); r += 1
stats_rows = [
    ("Media (μ)",       cv_acc.mean(), cv_f1.mean()),
    ("Desv. std (σ)",   cv_acc.std(),  cv_f1.std()),
    ("Mínimo",          cv_acc.min(),  cv_f1.min()),
    ("Máximo",          cv_acc.max(),  cv_f1.max()),
    ("IC 95% inferior", cv_acc.mean()-1.96*cv_acc.std(), cv_f1.mean()-1.96*cv_f1.std()),
    ("IC 95% superior", cv_acc.mean()+1.96*cv_acc.std(), cv_f1.mean()+1.96*cv_f1.std()),
]
for i, (label, a, f) in enumerate(stats_rows):
    _cell(ws5, r, 1, label, fill=FILL_ALT if i%2==0 else FILL_WHITE,
          font=Font(bold=True, name="Calibri", size=10), align=ALIGN_LEFT, border=BORDER_THIN)
    _cell(ws5, r, 2, round(a, 4), fill=FILL_ALT if i%2==0 else FILL_WHITE,
          font=FONT_NORMAL, align=ALIGN_CENTER, border=BORDER_THIN)
    _cell(ws5, r, 3, round(f, 4), fill=FILL_ALT if i%2==0 else FILL_WHITE,
          font=FONT_NORMAL, align=ALIGN_CENTER, border=BORDER_THIN)
    r += 1

r += 2
embed_image(ws5, p_cv,  f"A{r}", w=520, h=250); r += 18
embed_image(ws5, p_pca, f"A{r}", w=580, h=260)


# ──────────────────────────────────────────────────────────────────
# HOJA 6 : CONFIGURACIÓN
# ──────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("CONFIGURACIÓN")
ws6.sheet_view.showGridLines = False
set_col_widths(ws6, {1:30, 2:40})
title_row(ws6, 1, "CONFIGURACIÓN COMPLETA DEL MODELO Y EXPERIMENTO", 1, 2)

sections = {
    "MODELO SVM": [
        ("Algoritmo",              "Support Vector Classifier (SVC)"),
        ("Kernel",                 "RBF (Radial Basis Function)"),
        ("Parámetro C",            str(svm.C)),
        ("Gamma",                  str(svm.gamma)),
        ("Class weight",           "balanced"),
        ("Probability",            "True (Platt scaling para AUC)"),
        ("Decision function",      "ovr (one-vs-rest)"),
        ("Random state",           str(RANDOM_STATE)),
        ("Vectores de soporte",    str(svm.n_support_.sum())),
        ("SV por clase",           str(dict(zip(svm.classes_, svm.n_support_)))),
    ],
    "DATOS": [
        ("Dataset",                "HAM10000 Skin Lesion (ISIC 2018)"),
        ("Total muestras",         str(len(df))),
        ("Clases",                 ', '.join(classes)),
        ("Num. clases",            str(n_classes)),
        ("Test split",             f"{TEST_SIZE*100:.0f}%"),
        ("Train muestras",         str(len(y_train))),
        ("Test muestras",          str(len(y_test))),
        ("Train post-SMOTE",       str(len(y_train_res))),
        ("Distribución original",  str(dict(zip(*np.unique(y, return_counts=True))))),
    ],
    "PREPROCESAMIENTO": [
        ("Tamaño imagen",          f"{IMG_SIZE[0]}×{IMG_SIZE[1]} px"),
        ("Canales color",          "RGB (3 canales)"),
        ("Features originales",    str(int(X_train.shape[1]))),
        ("Normalización",          "StandardScaler (media=0, std=1)"),
        ("Reducción dim.",         "PCA"),
        ("Varianza PCA objetivo",  f"{PCA_VARIANCE*100:.0f}%"),
        ("Varianza PCA real",      f"{var_total*100:.2f}%"),
        ("Componentes PCA",        str(n_components)),
        ("Reducción features",     f"{reduction_pct:.1f}%"),
        ("Balanceo de clases",     "SMOTE (Synthetic Minority Oversampling)"),
    ],
    "TIEMPOS": [
        ("Carga imágenes",         f"{t_load:.2f} s"),
        ("Entrenamiento SVM",      f"{t_train:.2f} s"),
        ("Inferencia total",       f"{t_infer:.4f} s"),
        ("Inferencia/imagen",      f"{t_per_img_ms:.4f} ms"),
        ("Total pipeline",         f"{t_load+t_train+t_infer:.2f} s"),
    ],
    "ARCHIVOS GENERADOS": [
        ("Excel métricas",         str(EXCEL_PATH)),
        ("Modelo guardado",        str(MODEL_PATH)),
        ("JSON baseline",          str(JSON_PATH)),
        ("Directorio plots",       str(PLOT_DIR)),
    ],
}

r = 3
for section_name, items in sections.items():
    subheader_row(ws6, r, section_name, 1, 2); r += 1
    header_cells(ws6, r, ["Parámetro", "Valor"]); r += 1
    for i, (k, v) in enumerate(items):
        _cell(ws6, r, 1, k,
              fill=FILL_ALT if i%2==0 else FILL_WHITE,
              font=Font(bold=True, name="Calibri", size=10),
              align=ALIGN_LEFT, border=BORDER_THIN)
        _cell(ws6, r, 2, v,
              fill=FILL_ALT if i%2==0 else FILL_WHITE,
              font=FONT_NORMAL, align=ALIGN_LEFT, border=BORDER_THIN)
        r += 1
    r += 1


# ──────────────────────────────────────────────────────────────────
# HOJA 7 : COMPARACIÓN DE MODELOS
# ──────────────────────────────────────────────────────────────────
ws7 = wb.create_sheet("COMPARACIÓN_MODELOS")
ws7.sheet_view.showGridLines = False
comp_cols = [
    "Modelo","Fecha","Accuracy","Bal.Acc","F1 Macro","F1 Weighted",
    "AUC Macro","Kappa","MCC","Prec Macro","Rec Macro",
    "CV Acc (μ)","CV F1 (μ)","T.Train (s)","T.Inf/img (ms)","Notas"
]
set_col_widths(ws7, {i+1: 16 for i in range(len(comp_cols))})
ws7.column_dimensions['A'].width = 24
ws7.column_dimensions['B'].width = 20
ws7.column_dimensions[get_column_letter(len(comp_cols))].width = 28
title_row(ws7, 1,
          "COMPARACIÓN DE MODELOS — Línea base + Futuros modelos",
          1, len(comp_cols), height=30)

note = ws7.cell(row=2, column=1,
    value="ℹ️  La fila SVM (azul) es la LÍNEA BASE. "
          "Ejecuta append_metrics.py con cada nuevo modelo para añadir filas automáticamente.")
note.font = Font(italic=True, color="595959", name="Calibri", size=9)
ws7.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(comp_cols))

r = 4
header_cells(ws7, r, comp_cols)
r += 1

# Fila baseline SVM
baseline_vals = [
    MODEL_NAME, TIMESTAMP,
    round(acc,4), round(bal_acc,4), round(f1_macro,4), round(f1_weighted,4),
    round(auc_macro,4), round(kappa,4), round(mcc,4),
    round(prec_macro,4), round(rec_macro,4),
    round(cv_acc.mean(),4), round(cv_f1.mean(),4),
    round(t_train,2), round(t_per_img_ms,4), "★ LÍNEA BASE"
]
for c, val in enumerate(baseline_vals, 1):
    cell = ws7.cell(row=r, column=c, value=val)
    cell.fill = PatternFill("solid", fgColor="D9E1F2")
    cell.font = Font(bold=True, color="1F497D", name="Calibri", size=10)
    cell.alignment = ALIGN_CENTER; cell.border = BORDER_THIN
r += 1

# Filas placeholder para futuros modelos
future = [
    ("SVM_Optimizado",    "Ajuste hiperparámetros C, gamma (GridSearchCV)"),
    ("SVM_HOG",           "Features HOG en lugar de píxeles crudos"),
    ("Random Forest",     "Ensemble de árboles de decisión"),
    ("XGBoost",           "Gradient Boosting optimizado"),
    ("LightGBM",          "Gradient Boosting eficiente"),
    ("CNN_Simple",        "Red convolucional desde cero"),
    ("ResNet50",          "Transfer learning ImageNet — ResNet50"),
    ("EfficientNetB0",    "Transfer learning — EfficientNet B0"),
    ("VGG16",             "Transfer learning — VGG16"),
    ("DenseNet121",       "Transfer learning — DenseNet121"),
    ("ViT",               "Vision Transformer"),
]
for i, (model, nota) in enumerate(future):
    ws7.cell(row=r, column=1, value=model).font = Font(name="Calibri", size=10, color="7F7F7F")
    ws7.cell(row=r, column=len(comp_cols), value=nota).font = Font(italic=True, name="Calibri", size=9, color="7F7F7F")
    for c in range(1, len(comp_cols)+1):
        cell = ws7.cell(row=r, column=c)
        cell.fill = FILL_ALT if i%2==0 else FILL_WHITE
        cell.border = BORDER_THIN
        if c not in [1, len(comp_cols)]:
            cell.value = "—"
            cell.alignment = ALIGN_CENTER
            cell.font = Font(name="Calibri", size=10, color="BFBFBF")
    r += 1

# Guardar excel
wb.save(EXCEL_PATH)
print(f"  ✓ Excel guardado: {EXCEL_PATH}")


# ═════════════════════════════════════════════════════════════════════
# JSON — LÍNEA BASE (para append_metrics.py)
# ═════════════════════════════════════════════════════════════════════
baseline_json = {
    "run_id"        : RUN_ID,
    "model_name"    : MODEL_NAME,
    "timestamp"     : TIMESTAMP,
    "dataset"       : "HAM10000",
    "img_size"      : list(IMG_SIZE),
    "n_train"       : int(len(y_train)),
    "n_test"        : int(len(y_test)),
    "n_train_smote" : int(len(y_train_res)),
    "n_classes"     : int(n_classes),
    "classes"       : classes,
    "pca_components": int(n_components),
    "pca_variance"  : round(float(var_total), 6),
    "global_metrics": {
        "accuracy"        : round(float(acc),         6),
        "balanced_accuracy": round(float(bal_acc),    6),
        "f1_macro"        : round(float(f1_macro),    6),
        "f1_weighted"     : round(float(f1_weighted), 6),
        "f1_micro"        : round(float(f1_micro),    6),
        "precision_macro" : round(float(prec_macro),  6),
        "precision_weighted": round(float(prec_w),    6),
        "recall_macro"    : round(float(rec_macro),   6),
        "recall_weighted" : round(float(rec_w),       6),
        "auc_macro"       : round(float(auc_macro),   6),
        "auc_weighted"    : round(float(auc_weighted),6),
        "cohen_kappa"     : round(float(kappa),       6),
        "mcc"             : round(float(mcc),         6),
    },
    "per_class_metrics": {
        cls: {
            "precision": round(float(report_dict[cls]['precision']), 6),
            "recall"   : round(float(report_dict[cls]['recall']),    6),
            "f1_score" : round(float(report_dict[cls]['f1-score']),  6),
            "support"  : int(report_dict[cls]['support']),
            "auc"      : round(float(auc_per_class.get(cls, 0)), 6),
            "avg_prec" : round(float(ap_per_class.get(cls, 0)),  6),
        }
        for cls in classes
    },
    "cross_validation": {
        "n_folds"       : int(N_CV_FOLDS),
        "cv_acc_mean"   : round(float(cv_acc.mean()), 6),
        "cv_acc_std"    : round(float(cv_acc.std()),  6),
        "cv_acc_scores" : [round(float(v), 6) for v in cv_acc],
        "cv_f1_mean"    : round(float(cv_f1.mean()),  6),
        "cv_f1_std"     : round(float(cv_f1.std()),   6),
        "cv_f1_scores"  : [round(float(v), 6) for v in cv_f1],
    },
    "timing_seconds": {
        "load"       : round(float(t_load),    3),
        "train"      : round(float(t_train),   3),
        "infer_total": round(float(t_infer),   4),
        "infer_per_img_ms": round(float(t_per_img_ms), 4),
    },
    "model_config": {
        "kernel"    : svm.kernel,
        "C"         : svm.C,
        "gamma"     : str(svm.gamma),
        "probability": svm.probability,
        "n_support_total": int(svm.n_support_.sum()),
    },
}

with open(JSON_PATH, 'w', encoding='utf-8') as f:
    json.dump(baseline_json, f, indent=2, ensure_ascii=False)
print(f"  ✓ JSON baseline : {JSON_PATH}")


# ═════════════════════════════════════════════════════════════════════
# RESUMEN FINAL EN CONSOLA
# ═════════════════════════════════════════════════════════════════════
print("\n" + "=" * 65)
print("  RESULTADOS FINALES")
print("=" * 65)
print(f"  Accuracy          : {acc:.4f}")
print(f"  Balanced Accuracy : {bal_acc:.4f}")
print(f"  F1 Macro          : {f1_macro:.4f}")
print(f"  F1 Weighted       : {f1_weighted:.4f}")
print(f"  AUC Macro (OvR)   : {auc_macro:.4f}")
print(f"  Cohen's Kappa     : {kappa:.4f}")
print(f"  MCC               : {mcc:.4f}")
print(f"  CV Accuracy       : {cv_acc.mean():.4f} ± {cv_acc.std():.4f}")
print(f"  CV F1 Macro       : {cv_f1.mean():.4f} ± {cv_f1.std():.4f}")
print(f"  T. entrenamiento  : {t_train:.1f} s")
print(f"  T. inferencia/img : {t_per_img_ms:.3f} ms")
print("=" * 65)
print("\nARCHIVOS GENERADOS:")
print(f"  📊  {EXCEL_PATH}")
print(f"  🤖  {MODEL_PATH}")
print(f"  📋  {JSON_PATH}")
print(f"  📈  {PLOT_DIR}/  ({len(list(PLOT_DIR.glob('*.png')))} gráficos PNG)")
print("\nREPORTE POR CLASE:")
print(classification_report(y_test, y_pred, target_names=classes, zero_division=0))
print("✅ Pipeline completado.")
